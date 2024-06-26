#!/usr/bin/env python3
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QWidget
import os
import pandas as pd
import shutil




# Import dependencies
from os.path import join
from pandas import DataFrame, ExcelWriter
from sys import exit
global is_not_win
is_not_win = False

try:
    from include.meta.__init__ import *

except ModuleNotFoundError:
    print('Required files missing.\nExiting Application...')
    exit(-500)

try:
    import json
    import requests
    import cv2
    from PyQt5 import QtCore, QtGui, QtWidgets
    from functools import partial
    from pyzbar.pyzbar import decode
    from datetime import datetime, timedelta
    from pandas import DataFrame, ExcelWriter
    from smtplib import SMTP_SSL, SMTPAuthenticationError, SMTPServerDisconnected, SMTPRecipientsRefused
    from email.message import EmailMessage
    from socket import gaierror
    from qrcode import constants, QRCode
    from json import loads
    from hashlib import sha256
    from os.path import abspath, basename, dirname, join
    from math import floor
    from configparser import ConfigParser, NoOptionError
    from argparse import ArgumentParser

except ModuleNotFoundError:
    print("\nIncomplete installation. Install required pip packages.")
    print("List of required packages in 'requirements.txt' file.")
    print("\nRun following command to install required packages:")
    print("\n\npip install - r requirements.txt\n")
    exit(-11)

# Configure platform-specific values
try:
    # Only available on windows
    from winsound import Beep

except ModuleNotFoundError:
    # global is_not_win
    is_not_win = True


class Faculty:

    def __init__(self, filepath, token):

        # Set class-wide attributes
        # Assign parameters to class-wide variables
        self.token_size = token
        self.filepath = filepath

        # Create shell containers for class-wide use
        self.database = dict()
        self.session_faculty = dict()

        # Set static properties
        self.date = int(datetime.now().month + datetime.now().day + datetime.now().year)

        # Call methods in sequence
        self.read_db()
        self.generate_sessions()

    # Read database file to shell container
    def read_db(self):

        try:
            with open(self.filepath, mode='r') as faculty_file:
                self.database = loads(faculty_file.read())
                faculty_file.close()
        except FileNotFoundError:
            print('Required resources not complete! Check requirements.')
            exit(-404)

    # Generate sessions for present day
    def generate_sessions(self):

        # Hashing algorithm
        for faculty in self.database:

            # Mangle numerical part of faculty id with string part after having multiplied former with present date
            mangler = faculty['Code'][:3] + str(int(faculty['Code'][3:]) * self.date)

            # Stable hash creates instance of sha256() - onto hashing function
            stable_hash = sha256()

            # Pass mangler as c-string
            stable_hash.update(mangler.encode())

            # Generate hash-bytes
            hash_bytes = stable_hash.digest()

            # Convert hash bytes to integer; limit them by stable modulus hashing against token_size integer
            faculty['session'] = int.from_bytes(hash_bytes, byteorder='big', signed=False) % self.token_size

    # Authenticate session input
    def auth(self, token):

        # Check length of token against length of token_size(known); check if token is all digits
        if len(token) <= len(str(self.token_size)) and token.isdigit():

            # Check is token is a sub-string of database - optimization
            if str(token) in str(self.database):
                faculty = dict()

                # Fetch first result whose hash matches with token
                for faculty_member in self.database:
                    if int(token) == int(faculty_member['session']):
                        faculty = faculty_member
                        break

                # Assign session faculty in first iteration only
                if not str(faculty) == str(self.session_faculty):
                    self.session_faculty = faculty.copy()
                    return self.session_faculty


class Student:

    def __init__(self, filepath, output_dir, console_output):
        # Set class-wide attributes
        # Assign parameters to class-wide variables
        self.data_file = filepath
        self.output_dir = output_dir

        # Function bound to Application class; prints message to emulated console
        self.console_output = console_output

        # Create shell containers for class-wide use
        self.database = list()
        self.student_list = list()
        self.attended_students = set()  # Set to store attended students

        # Call method in sequence
        self.read_db()

    # Read database file to shell container; sort list
    def read_db(self):
        try:
            with open(self.data_file, mode='r') as database:
                self.database = loads(database.read())
                database.close()
        except FileNotFoundError:
            print('Required resources not complete! Check requirements.')
            exit(-404)

        # Sort database by roll_number in increasing order
        self.database = sorted(self.database, key=lambda i: i['Roll_Number'])
        #Input your PhilSMS API key here
        self.api_key = ""

    # Validate attendee entry against database - uses binary search algorithm
    def validate(self, roll, name, base_list=None, left_index=0, right_index=None):
        # Case when called from outer scope
        if not base_list and not left_index and not right_index:
            base_list = self.database
            right_index = len(self.database) - 1

        # Check base case
        if right_index >= left_index:
            # Find middle of array length and floor it down to integer
            mid = floor((left_index + right_index) / 2)

            # If element is present at the middle itself
            try:
                if int(base_list[mid]['Roll_Number']) == int(roll) and str(base_list[mid]['Name']) == str(name):
                    # Check if student has already attended
                    if roll not in self.attended_students:
                        # Mark student as attended
                        self.attended_students.add(roll)
                        # Send SMS when attendee is validated   
                        mobile_number = str(base_list[mid]['Mobile_Number'])  # Convert to string
                        message = f"Dear Guardian/Parent, \n This message is to notify you that your child has successfully reached the STI Calbayog premises at {self.get_current_datetime()}"
                        self.send_sms(mobile_number, message)
                    return base_list[mid]


                    
                # If element is smaller than mid, check left_index subarray
                elif int(base_list[mid]['Roll_Number']) > int(roll):
                    return self.validate(base_list=base_list, left_index=left_index, right_index=mid - 1,
                                         roll=roll, name=name)

                # Else check     right subarray
                else:
                    return self.validate(base_list=base_list, left_index=mid + 1, right_index=right_index,
                                         roll=roll, name=name)

            except ValueError:
                return None

        else:
            # Element is not present in the array
            return None

    # Function to send SMS using PhilSMS API
    def send_sms(self, mobile_number, message):
        url = "https://app.philsms.com/api/v3/sms/send"
        payload = {
            "recipient": mobile_number,
            "sender_id": "PhilSMS",  # Replace with your sender ID
            "message": message,
            # Optional parameters
            # "schedule_time": "2024-03-21 12:00:00",
            # "dlt_template_id": "YOUR_DLT_TEMPLATE_ID"
        }
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        try:
            response = requests.post(url, data=json.dumps(payload), headers=headers)
            if response.status_code == 200:
                print("SMS sent successfully!")
            else:
                print("Failed to send SMS! Status code:", response.status_code)
        except Exception as e:
            print("An error occurred while sending SMS:", e)

    # Helper function to get current date and time
    def get_current_datetime(self):
        # You can customize the format as per your requirement
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Load student data from JSON file
    with open('C:/Users/Sample file path directory/PaSCAN/resource/student.json') as f:
        student_data = json.load(f)


    # Generate attendee QR code
    def code_generator(self, main_application):

        # Generate QR code for each attendee in dataset
        for each_attendee in self.database:

            # Print attendee data
            self.console_output(list(each_attendee.values()))

            # Set QR code instance properties and save it to output_dir folder
            qr = QRCode(
                version=1,
                error_correction=constants.ERROR_CORRECT_H,
                box_size=10,
                border=4,
            )
            qr.add_data(list(each_attendee.values()))
            qr.make(fit=True)

            img = qr.make_image(fill_color="black", back_color="white")

            img.save(join(self.output_dir, each_attendee['Roll_Number'] + ' - ' +
                          each_attendee['Name'] + '.png'), 'PNG')

            # Unfreeze event loop - tell it to process events
            main_application.processEvents()

class Logout:

    def __init__(self, filepath, output_dir, console_output):
        # Set class-wide attributes
        # Assign parameters to class-wide variables
        self.data_file = filepath
        self.output_dir = output_dir

        # Function bound to Application class; prints message to emulated console
        self.console_output = console_output

        # Create shell containers for class-wide use
        self.database = list()
        self.student_list = list()
        self.attended_students = set()  # Set to store attended students

        # Call method in sequence
        self.read_db()

    # Read database file to shell container; sort list
    def read_db(self):
        try:
            with open(self.data_file, mode='r') as database:
                self.database = loads(database.read())
                database.close()
        except FileNotFoundError:
            print('Required resources not complete! Check requirements.')
            exit(-404)

        # Sort database by roll_number in increasing order
        self.database = sorted(self.database, key=lambda i: i['Roll_Number'])
        #Input your PhilSMS API Key here
        self.api_key = ""

    # Validate attendee entry against database - uses binary search algorithm
    def validate(self, roll, name, base_list=None, left_index=0, right_index=None):
        # Case when called from outer scope
        if not base_list and not left_index and not right_index:
            base_list = self.database
            right_index = len(self.database) - 1

        # Check base case
        if right_index >= left_index:
            # Find middle of array length and floor it down to integer
            mid = floor((left_index + right_index) / 2)

            # If element is present at the middle itself
            try:
                if int(base_list[mid]['Roll_Number']) == int(roll) and str(base_list[mid]['Name']) == str(name):
                    # Check if student has already attended
                    if roll not in self.attended_students:
                        # Mark student as attended
                        self.attended_students.add(roll)
                        # Send SMS when attendee is validated   
                        mobile_number = str(base_list[mid]['Mobile_Number'])  # Convert to string
                        logoutmessage = f"Dear Guardian/Parent, \n This message is to notify you that your child has departed from STI Calbayog premises at {self.get_current_datetime()}"
                        self.logoutsend_sms(mobile_number, logoutmessage)
                    return base_list[mid]

                # If element is smaller than mid, check left_index subarray
                elif int(base_list[mid]['Roll_Number']) > int(roll):
                    return self.validate(base_list=base_list, left_index=left_index, right_index=mid - 1,
                                         roll=roll, name=name)

                # Else check right subarray
                else:
                    return self.validate(base_list=base_list, left_index=mid + 1, right_index=right_index,
                                         roll=roll, name=name)

            except ValueError:
                return None

        else:
            # Element is not present in the array
            return None

    # Function to send SMS using PhilSMS API
    def logoutsend_sms(self, mobile_number, logoutmessage):
        url = "https://app.philsms.com/api/v3/sms/send"
        payload = {
            "recipient": mobile_number,
            "sender_id": "PhilSMS",  # Replace with your sender ID
            "message": logoutmessage,
        }
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        try:
            response = requests.post(url, data=json.dumps(payload), headers=headers)
            if response.status_code == 200:
                print("SMS sent successfully!")
            else:
                print("Failed to send SMS! Status code:", response.status_code)
        except Exception as e:
            print("An error occurred while sending SMS:", e)

    # Helper function to get current date and time
    def get_current_datetime(self):
        # You can customize the format as per your requirement
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Load student data from JSON file
    with open('C:/Users/Sample file path directory/PaSCAN/resource/student.json') as f:
        student_data = json.load(f)

class Token:

    def __init__(self, faculty_path, output_dir, token_size, mailer_object, console_output):

        # Set class-wide attributes
        # Assign parameters to class-wide variables
        self.token_size = token_size
        self.faculty_path = faculty_path
        self.output_dir = output_dir

        # Function bound to Application class; prints message to emulated console
        self.console_output = console_output

        # Create shell containers for class-wide use
        self.database = list()

        # Fetch mailer object to enable mailing capabilities
        self.mailer = mailer_object

        # Set static properties
        self.date = int(datetime.now().month + datetime.now().day + datetime.now().year)

        # Call methods in sequence
        self.read_db()

    # Read database file to shell container
    def read_db(self):
        with open(self.faculty_path, mode='r') as database:
            self.database = loads(database.read())
            database.close()

    # Generate sessions for present day; generate QR codes session token; mail QR codes to respective faculty
    def generate_session(self, main_application):

        # Iterate for each faculty entry in self.database:
        for faculty in self.database:

            # Mangle numerical part of faculty id with string part after having multiplied former with present date
            mangler = (faculty['Code'])[:3] + str(int(faculty['Code'][3:]) * self.date)

            # Stable hash creates instance of sha256() - onto hashing function
            stable_hash = sha256()

            # Pass mangler as c-string
            stable_hash.update(mangler.encode())

            # Generate hash-bytes
            hash_bytes = stable_hash.digest()

            # Convert hash bytes to integer; limit them by stable modulus hashing against token_size integer
            faculty['session'] = int.from_bytes(hash_bytes, byteorder='big', signed=False) % self.token_size

            # Print faculty data - without the actual tokens
            self.console_output(list(faculty.values())[0:-1])

            # Set QR code instance properties and save it to output_dir folder
            qr = QRCode(version=1, box_size=10, border=4,
                        error_correction=constants.ERROR_CORRECT_H,)
            qr.add_data(faculty['session'])
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")

            # Name of token file (image file, png extension) and path
            token_name = faculty['Code'] + '_' + faculty['Name'].replace(' ', '_') + '.png'
            img_path = join(self.output_dir, token_name)

            # Save token file in PNG format
            img.save(img_path, 'PNG')

            # Mail token QR code to respective faculty
            self.mailer.send_token(email=faculty['Email'], attachment=img_path, name=faculty['Name'])

            # Unfreeze event loop - tell it to process events
            main_application.processEvents()


class Timer:

    def __init__(self, filepath):

        # Read lecture breakpoints to a container dict()
        try:
            with open(filepath, mode='r') as breakpoints:
                self.lecture = loads(breakpoints.read())
                breakpoints.close()
        except FileNotFoundError:
            print('Required resources not complete! Check requirements.')
            exit(-404)

        # Read key name from timing.json file
        key_name = list(self.lecture.keys())[0]

        # Create list of time slots as tuple pairs from timing breakpoints
        self.timing_list = [(i, j) for i, j in zip(self.lecture[key_name][:-1],
                                                   self.lecture[key_name][1:])]

    # Determine time slot of current lecture
    def lecture_time(self):

        present = (datetime.now().year, datetime.now().month, datetime.now().day)

        # Iterate over each time slot to determine if it encompasses current time
        for entry in self.timing_list:

            # Process time slot format for timing info
            start_time, end_time = entry[0], entry[1]
            start_hour = int(start_time[:start_time.index(':')])
            start_minute = int(start_time[(start_time.index(':') + 1):])
            end_hour = int(end_time[:end_time.index(':')])
            end_minute = int(end_time[(end_time.index(':')+1):])

            start = datetime(year=present[0], month=present[1], day=present[2], hour=start_hour, minute=start_minute)
            try:
                # Attempt to create the datetime object
                end = datetime(year=present[0], month=present[1], day=present[2], hour=end_hour, minute=end_minute)
            except ValueError as e:
                # Handle the error or print the error message
                print("Error creating datetime object:", e)

            # Check the value of end_hour
            if end_hour < 0 or end_hour > 23:
                # Handle the case where end_hour is outside the valid range
                print("Error: end_hour must be in the range 0 to 23.")


            # Return matching time slot along with lecture end time
            if start < datetime.now() < end:
                return datetime(present[0], present[1], present[2],
                                hour=end_hour, minute=end_minute), self.timing_list.index(entry), entry

        # Else return None if loop runs through


class Scheduler:

    def __init__(self, batch_name, output_dir_path, path_timing, path_lecture, console_output):

        # Set class-wide attributes
        # Assign parameters to class-wide variables
        self.date = str(datetime.now().year) + '{:02d}'.format(datetime.now().month) +\
                    '{:02d}'.format(datetime.now().day)

        # Assign name of excel file (xlsx extension)
        self.filename = batch_name + ' - ' + self.date + '.xlsx'
        self.filepath = join(output_dir_path, self.filename)

        self.path_timing = path_timing
        self.path_lecture = path_lecture

        self.temp_key = 'start'

        # Function bound to Application class; prints message to emulated console
        self.console_output = console_output

        # Create shell containers for class-wide use
        self.timing = dict()
        self.lecture_table = dict()

        # Call methods in sequence
        self.structure()

    # Read timing and lecture files; create timing structure
    def structure(self):
        with open(self.path_timing, mode='r') as time_points:
            self.timing = loads(time_points.read())
            time_points.close()

        # Read key name from timing.json file
        key_name = list(self.timing.keys())[0]

        # Create list of time slots as tuple pairs from timing breakpoints and add to self.timing dictionary
        self.timing[self.temp_key] = [(i, j) for i, j in zip(self.timing[key_name][:-1],
                                                             self.timing[key_name][1:])]

        try:
            with open(self.path_lecture, mode='r') as table:
                self.lecture_table = loads(table.read())
                table.close()
        except FileNotFoundError:
            print('Required resources not complete! Check requirements.')
            exit(-404)

    # Return subject name from lecture table
    def lecture(self, section, index):

        # Fetch lecture table only for the given section
        lecture_schedule = self.lecture_table[section]

        # Return subject name index present day with index lecture number
        try:
            return lecture_schedule[list(lecture_schedule.keys())[datetime.today().weekday()]][index]
        except IndexError:
            return None

    # Generate lecture schedule excel file with multiple sheets for different sections
    def schedule(self):

        # Print message to let user know of this action
        self.console_output('Generating lecture schedule (Excel file: ' + self.filename + ')')

        # Instantiate writer engine - with path to export it to
        writer = ExcelWriter(self.filepath, engine='xlsxwriter')

        # Instantiate workbook object - bound to writer object
        workbook = writer.book

        # Add formatting options
        # bold = workbook.add_format({'bold': True, 'center_across': True})
        center = workbook.add_format({'center_across': True})

        # Iterate over lectures dict() - may contain multiple lecture table record for different sections
        for key, val in self.lecture_table.items():

            # Key contains section name, val contains section schedule
            # Create dataframe
            data_frame = DataFrame(val.values(), index=val.keys(),
                                   columns=[str(lecture_time).strip(')(').replace("'", "").replace(",", " -") for
                                            lecture_time in self.timing[self.temp_key]])

            # Add a worksheet to workbook - key is the name of the sheet
            worksheet = workbook.add_worksheet(key)
            writer.sheets[key] = worksheet

            # Format column width
            worksheet.set_column(0, (len(self.timing[self.temp_key])), 35, center)

            # Add dataframe as a single sheet using writer engine
            data_frame.to_excel(writer, sheet_name=key, index_label=key)

        # Save writer file to the path passed as a parameter earlier
        writer._save()

class LogExport:
    def __init__(self, folder, name):

        # Set class-wide attributes
        # Assign parameters to class-wide variables
        self.folder = folder
        self.name = name

        # Set static properties
        self.present = str(datetime.now().year) + '-' + str(datetime.now().month) + '-' + str(datetime.now().day)
        self.dateCreated = 'Logout' 

        # Add date to filename to keep one file record for the entire day
        self.file = self.name + ' ' + self.present + ' ' + self.dateCreated + '.xlsx'
        self.path = join(self.folder, self.file)



    def get_current_datetime_forExcel(self):
        # You can customize the format as per your requirement %Y-%m-%d
        return datetime.now().strftime("%B %d %Y %H:%M:%S")


    # Create and export dataframe from containers, to excel file(s)
    def export(self, attendance):

        # Remove the function call to path_lecture

        # Instantiate writer engine - with path to export it to
        writer = ExcelWriter(self.path, engine='xlsxwriter')

        # Instantiate workbook object - bound to writer object
        workbook = writer.book

        # Add formatting options
        bold = workbook.add_format({'bold': True, 'center_across': True})
        # center = workbook.add_format({'center_across': True})
        center = workbook.add_format({'center_across': False, 'align': 'center'})

        # Iterate over attendance dict() - may contain multiple attendance record of entire day
        for key, val in attendance.items():

            # Key contains name of subject + time slot; perform segregation
            lecture_time = key[key.index('('):]

            key = key[:key.index('(')-1]
            # Make key unique by adding timestamp to subject name
            key = key[:25] + ' ' + lecture_time[1:lecture_time.index('-')].replace(':', '')

            # Create host and attendees dataframes
            host_df = DataFrame([val[list(val.keys())[0]]['Name'].title()],
                                index=[val[list(val.keys())[0]]['Code']],
                                columns=['Host Faculty'])
            attendee_df = DataFrame([attendee['Name'].title() for attendee in val[list(val.keys())[1]]],
                                    index=[attendee['Roll_Number'] for attendee in val[list(val.keys())[1]]],
                                    columns=['Attendees']).sort_index()

            # Add a worksheet to workbook - key is the name of the sheet
            worksheet = workbook.add_worksheet(str(key))
            writer.sheets[str(key)] = worksheet
            worksheet.write_string(0, 0, lecture_time, bold)
            worksheet.write_string(0, 1, key[:-5], bold)

            # Add a new column for the specified time the students attended the school
            worksheet.write_string(0, 2, 'Time Departed', bold)

            # Format column width
            worksheet.set_column(0, 2, 35, center)

            # Add dataframes to a single sheet using writer engine
            host_df.to_excel(writer, sheet_name=str(key), index_label='Code', startrow=1, startcol=0)
            attendee_df.to_excel(writer, sheet_name=str(key), index_label='Roll Number',
                                 startrow=host_df.shape[0] + 3, startcol=0)
            
            # Center the text in the second column
            worksheet.set_column(1, 1, 35, center)

            # Get the current datetime
            current_datetime = self.get_current_datetime_forExcel()

            # Write the current datetime to the new column for each attendee
            for i, attendee in enumerate(val[list(val.keys())[1]]):
                worksheet.write(i + host_df.shape[0] + 4, 2, current_datetime)


        # Remove the position for "0,1" in the Excel file
        worksheet = writer.sheets[str(key)]
        worksheet.write_string(0, 1, 'Left The School Premises', bold)  # This will write an empty string to the position "0,1"

        # Save writer file to the path passed as a parameter earlier
        writer._save()

        # Return export filepath
        return self.path

class Export:

    def __init__(self, folder, name):

        # Set class-wide attributes
        # Assign parameters to class-wide variables
        self.folder = folder
        self.name = name

        # Set static properties
        self.present = str(datetime.now().year) + '-' + str(datetime.now().month) + '-' + str(datetime.now().day)
        self.dateCreated = 'LogIn'

        # Add date to filename to keep one file record for the entire day
        self.file = self.name + ' ' + self.present + ' ' + self.dateCreated + '.xlsx'
        self.path = join(self.folder, self.file)



    def get_current_datetime_forExcel(self):
        # You can customize the format as per your requirement %Y-%m-%d
        return datetime.now().strftime("%B %d %Y %H:%M:%S")


    # Create and export dataframe from containers, to excel file(s)
    def export(self, attendance):

        # Remove the function call to path_lecture

        # Instantiate writer engine - with path to export it to
        writer = ExcelWriter(self.path, engine='xlsxwriter')

        # Instantiate workbook object - bound to writer object
        workbook = writer.book

        # Add formatting options
        bold = workbook.add_format({'bold': True, 'center_across': True})
        # center = workbook.add_format({'center_across': True})
        center = workbook.add_format({'center_across': False, 'align': 'center'})

        # Iterate over attendance dict() - may contain multiple attendance record of entire day
        for key, val in attendance.items():

            # Key contains name of subject + time slot; perform segregation
            lecture_time = key[key.index('('):]

            key = key[:key.index('(')-1]
            # Make key unique by adding timestamp to subject name
            key = key[:25] + ' ' + lecture_time[1:lecture_time.index('-')].replace(':', '')

            # Create host and attendees dataframes
            host_df = DataFrame([val[list(val.keys())[0]]['Name'].title()],
                                index=[val[list(val.keys())[0]]['Code']],
                                columns=['Host Faculty'])
            attendee_df = DataFrame([attendee['Name'].title() for attendee in val[list(val.keys())[1]]],
                                    index=[attendee['Roll_Number'] for attendee in val[list(val.keys())[1]]],
                                    columns=['Attendees']).sort_index()

            # Add a worksheet to workbook - key is the name of the sheet
            worksheet = workbook.add_worksheet(str(key))
            writer.sheets[str(key)] = worksheet
            worksheet.write_string(0, 0, lecture_time, bold)
            worksheet.write_string(0, 1, key[:-5], bold)

            # Add a new column for the specified time the students attended the school
            worksheet.write_string(0, 2, 'Time Attended', bold)

            # Format column width
            worksheet.set_column(0, 2, 35, center)

            # Add dataframes to a single sheet using writer engine
            host_df.to_excel(writer, sheet_name=str(key), index_label='Code', startrow=1, startcol=0)
            attendee_df.to_excel(writer, sheet_name=str(key), index_label='Roll Number',
                                 startrow=host_df.shape[0] + 3, startcol=0)
            
            # Center the text in the second column
            worksheet.set_column(1, 1, 35, center)

            # Get the current datetime
            current_datetime = self.get_current_datetime_forExcel()

            # Write the current datetime to the new column for each attendee
            for i, attendee in enumerate(val[list(val.keys())[1]]):
                worksheet.write(i + host_df.shape[0] + 4, 2, current_datetime)


        # Remove the position for "0,1" in the Excel file
        worksheet = writer.sheets[str(key)]
        worksheet.write_string(0, 1, 'Arrived At School Premises', bold)  # This will write an empty string to the position "0,1"

        # Save writer file to the path passed as a parameter earlier
        writer._save()

        # Return export filepath
        return self.path


class Mailer:

    def __init__(self, batch, email, password, hod_email, console_output, main_window):

        # Set class-wide attributes
        # Assign parameters to class-wide variables
        self.email = email
        self.password = password
        self.hod_email = hod_email
        self.batch = batch

        # Function bound to Application class; prints message to emulated console
        self.console_output = console_output

        # Instance of main application - pass control of the event loop
        self.main_window = main_window

    # Create mail-able message containing attendance information
    def send_attendance(self, attachment, attendees_len=None, lecture=None, email=None, lecture_len=None):

        # Instantiate EmailMessage object
        msg = EmailMessage()
        msg['From'] = self.email

        # Define case for mailing to host faculty
        if attendees_len and lecture and email and not lecture_len:
            msg['Subject'] = f'Lecture Attendance: {lecture}'
            msg['To'] = email
            msg.set_content(f'{attendees_len} attendees.\n\nDate: {datetime.today().date()} '
                            f'({datetime.now().strftime("%A")})\nLecture: {lecture}')
            msg.add_alternative(f'<strong>{attendees_len} attendees.</strong><br><br>Date: {datetime.today().date()} '
                                f'({datetime.now().strftime("%A")})<br>Lecture: {lecture}', subtype='html')

        # Define case for mailing to HOD
        elif lecture_len:
            msg['Subject'] = f'Attendance {self.batch} {datetime.today().date()}'
            msg['To'] = self.hod_email
            msg.set_content(f'{lecture_len} lectures held today.\n\nDate: {datetime.today().date()} '
                            f'({datetime.now().strftime("%A")}).')
            msg.add_alternative(f'<strong>{lecture_len} lectures held today.</strong><br><br>Date: '
                                f'{datetime.today().date()} ({datetime.now().strftime("%A")}).', subtype='html')

        # Read attachment file from file system
        with open(attachment, mode='rb') as record:
            record_data = record.read()
            record_name = basename(record.name)

        # Add attachment to message
        msg.add_attachment(record_data, maintype='application', subtype='octet-stream', filename=record_name)

        # Pass message to send mail
        self.send(msg)

    # Create mail-able message containing session token
    def send_token(self, attachment, email=None, name=None):

        # Instantiate EmailMessage object
        msg = EmailMessage()
        msg['From'] = self.email

        # Makes sure name and email of the faculty are given - deprecated other case where only name is given
        if email and name:
            msg['Subject'] = f'Access Token: {name}'
            msg['To'] = email
            msg.set_content(f'Valid for: {datetime.today().date()} ({datetime.now().strftime("%A")})')
            msg.add_alternative(f'Valid for: <strong>{datetime.today().date()}</strong> '
                                f'({datetime.now().strftime("%A")})', subtype='html')

        # Read attachment file from file system
        with open(attachment, mode='rb') as record:
            record_data = record.read()
            record_name = basename(record.name)

        # Add attachment to message
        msg.add_attachment(record_data, maintype='image', subtype='png', filename=record_name)

        # Pass message to send mail function
        self.send(msg)

    # Send mail
    def send(self, message):

        # Try sending mail
        try:
            with SMTP_SSL('smtp.gmail.com', 465) as smtp:

                # Login to controller account; print message
                self.console_output('Sending mail now!')
                smtp.login(self.email, self.password)

                # Unfreeze event loop - network operation
                self.main_window.processEvents()

                # Send message
                smtp.send_message(message)

        # Except cases for connection error, socket busy, or authentication failure - print message
        except SMTPAuthenticationError:
            self.console_output('Authentication Failure!\nCheck credentials in Configuration section.\n'
                                'Also check support for less secure apps in your account:\n'
                                'Login to controller account.\nGoto: myaccount.google.com/lesssecureapps\n'
                                'Allow less secure apps access.\n\nToken saved locally instead')

        except (gaierror, SMTPServerDisconnected, SMTPRecipientsRefused, OSError) as connectionError:
            self.console_output('Failure to establish connection with the server.\nToken saved locally instead.')


class Config:
    def __init__(self, qicon, obj, console_output):

        # Set class-wide attributes
        # Assign parameters to class-wide variables
        self.obj = obj
        self.console_output = console_output

        self.qicon = qicon

        self.qtranslate = self.qtranslate = QtCore.QCoreApplication.translate
        self.centralwidget_name = 'PaSCAN Dashboard'

    # Launch QDialog box to configure PaSCAN values
    def config_manager(self):

        # Instantiate QDialog() object and set it's properties
        config_dialog = QtWidgets.QDialog()
        config_dialog.setObjectName(self.centralwidget_name)
        config_dialog.setFixedSize(544, 346)
        config_dialog.setWindowTitle(self.qtranslate(self.centralwidget_name, "PaSCAN Configure"))
        config_dialog.setWindowIcon(self.qicon)

        # Instantiate QPushButton() object as Save button
        self.save_button = QtWidgets.QPushButton(config_dialog)

        # Read config items from config value(s) imported earlier
        self.configurations = self.obj.attribute.config.items(self.obj.attribute.config.sections()[0])

        # Create records and reverse-lookup records for future use
        self.key_name = {'token_limit': 'Token Limit: ',
                         'warning_period_minutes': 'Warning Time (Minutes): ',
                         'batch_name': 'Batch/Section Name: ',
                         'hod_email': 'Department Central Email: ',
                         'PaSCAN_email': 'Controller Email: ',
                         'PaSCAN_password': 'Controller Password: '}

        self.val_name = {v: k for k, v in self.key_name.items()}

        # Keep count on function calls
        self.count = 0

        # Length values for element placement on QDialog box
        _txtbox_offset_x_ = 30
        _txtbox_height_ = 25
        _txtbox_offset_margin_ = 15

        # Dynamically instantiate QLabels and QLineText fields as per items in config container
        for element in self.configurations:
            setattr(self, 'label_' + element[0], QtWidgets.QLabel(config_dialog))
            label = getattr(self, 'label_' + element[0])
            label.setGeometry(QtCore.QRect(40, _txtbox_offset_x_, 225, _txtbox_height_))
            # label.setFrameShape(QtWidgets.QFrame.Box)
            label.setText(self.qtranslate(self.centralwidget_name, self.key_name[element[0]]))
            label.setAlignment(QtCore.Qt.AlignCenter)

            setattr(self, 'txtbox_' + element[0], QtWidgets.QLineEdit(config_dialog))
            txt_box = getattr(self, 'txtbox_' + element[0])
            txt_box.setGeometry(QtCore.QRect(275, _txtbox_offset_x_, 225, _txtbox_height_))
            txt_box.setText(self.qtranslate(self.centralwidget_name, element[1]))
            txt_box.setAlignment(QtCore.Qt.AlignCenter)

            # Update offset value for next set of placements
            _txtbox_offset_x_ = _txtbox_offset_x_ + _txtbox_height_ + _txtbox_offset_margin_

            # Define click slot to connect to save_config() function
            self.save_button.clicked.connect(partial(self.save_config, label, txt_box))

        # Translate save_button properties
        self.save_button.setGeometry(
            QtCore.QRect(40, _txtbox_offset_x_+10, 460, int(_txtbox_height_ * 1.5)))
        self.save_button.setText(self.qtranslate(self.centralwidget_name, 'Save Configuration'))

        # Show dialog box
        config_dialog.show()
        config_dialog.exec_()

    # Function to save updated configuration
    def save_config(self, qlabel, qtxtbox):

        # Set values to config instance
        self.obj.attribute.config.set(self.obj.attribute.config.sections()[0],
                                      self.val_name[qlabel.text()], qtxtbox.text())

        # Update count - this function gets called mutiple times for a single click due to multiple slots
        # Save only once when all calls are executed; update count
        self.count = self.count + 1
        if self.count == len(self.configurations):
            # Save configuration container to a file
            with open(self.obj.path_config, mode='w') as config_file:
                self.obj.attribute.config.write(config_file)
                config_file.close()

            # Print message and update count
            self.console_output('Configuration saved! Restart PaSCAN to load new settings.')
            self.count = 0


class Attribute:

    # Meta class
    # Pre-set attributes that define behaviour of application; Used by all other classes
    def __init__(self, path_config):

        # Instantiate ConfigParser class; read config.ini file
        self.config = ConfigParser()
        self.config.read(path_config)

        # Import values from configuration file
        try:
            _index_name_ = self.config.sections()[0]
        except IndexError:
            print('Configuration file does not exist, or is empty! Application couldn\'t start.')
            exit(-100)

        try:
            self.tokenLimit = self.config.getint(_index_name_, 'token_limit')
            self.warning_period_minutes = self.config.getint(_index_name_, 'warning_period_minutes')
            self.batch_name = self.config.get(_index_name_, 'batch_name')
            self.hod_email = self.config.get(_index_name_, 'hod_email')
            self.PaSCAN_email = self.config.get(_index_name_, 'PaSCAN_email')
            self.PaSCAN_password = self.config.get(_index_name_, 'PaSCAN_password')

        except NoOptionError:
            print('Improper configuration file! Application couldn\'t start.')
            exit(-100)

        self.isAuthenticated = False
        self.isWarned = False
        self.isFlushed = False
        self.host_faculty = dict()
        self.attendees = list()
        self.attendance_all = dict()
        self.lecture_time = list()
        self.lecture_number = int()


class Object:

    # Meta class
    # Instantiate objects as utilities needed by other classes
    def __init__(self, console_output, qicon, application_window):

        # Meta - assign folder names containing other folders or file; fetch their absolute path
        self.json_folder_name = 'resource'
        self.database_folder_name = 'database'
        self.config_folder_name = 'config'

        self.json_folder_path = abspath(join(dirname(__file__), self.json_folder_name))
        self.database_folder_path = abspath(join(dirname(__file__), self.database_folder_name))
        self.config_folder_path = abspath(join(dirname(__file__), self.config_folder_name))

        # Assign filename containing operational info; fetch their path - uses meta
        self.file_faculty = 'faculty.json'
        self.file_student = 'student.json'
        self.file_lecture = 'lecture.json'
        self.file_timing = 'timing.json'

        self.path_faculty = join(self.json_folder_path, self.file_faculty)
        self.path_student = join(self.json_folder_path, self.file_student)
        self.path_lecture = join(self.json_folder_path, self.file_lecture)
        self.path_timing = join(self.json_folder_path, self.file_timing)

        # Assign folders to export to; fetch path - uses meta
        self.token_folder_name = 'session'
        self.attendee_folder_name = 'attendees'
        self.schedule_folder_name = 'schedule'
        self.attendance_folder_name = 'attendance'

        self.token_folder_path = join(self.database_folder_path, self.token_folder_name)
        self.attendee_folder_path = join(self.database_folder_path, self.attendee_folder_name)
        self.schedule_folder_path = join(self.database_folder_path, self.schedule_folder_name)
        self.attendance_folder_path = join(self.database_folder_path, self.attendance_folder_name)

        # Assign configuration file
        self.file_config = 'config.ini'
        self.path_config = join(self.config_folder_path, self.file_config)

        # Instantiate objects
        # Create instance of Attribute class to fetch attributes from
        # Import console_output instance bound to Application application which renders output message to
        # emulated console - this needs to be passed as an argument to other classes
        self.attribute = Attribute(self.path_config)
        self.console_output = console_output

        # Instantiate objects using __init__() and attribute object values
        self.faculty = Faculty(filepath=self.path_faculty, token=self.attribute.tokenLimit)
        self.student = Student(filepath=self.path_student, output_dir=self.attendee_folder_path,
                               console_output=self.console_output)
        self.logout = Logout(filepath=self.path_student, output_dir=self.attendee_folder_path,
                               console_output=self.console_output)
        self.export = Export(folder=self.attendance_folder_path, name=self.attribute.batch_name)
        self.logexport = LogExport(folder=self.attendance_folder_path, name=self.attribute.batch_name)
        self.timer = Timer(filepath=self.path_timing)
        self.scheduler = Scheduler(batch_name=self.attribute.batch_name, output_dir_path=self.schedule_folder_path,
                                   path_lecture=self.path_lecture, path_timing=self.path_timing,
                                   console_output=self.console_output)
        self.mailer = Mailer(batch=self.attribute.batch_name, email=self.attribute.PaSCAN_email,
                             password=self.attribute.PaSCAN_password, hod_email=self.attribute.hod_email,
                             console_output=self.console_output, main_window=application_window)
        self.token = Token(faculty_path=self.path_faculty, output_dir=self.token_folder_path,
                           token_size=self.attribute.tokenLimit, mailer_object=self.mailer,
                           console_output=self.console_output)
        self.config = Config(obj=self, console_output=self.console_output, qicon=qicon)

    # Function when called returns instance of attribute object used in this class
    def return_attribute_obj(self):
        return self.attribute


class Utility:

    # Non-visual feedback - audio
    def beep(self, frequency=2500, duration=300):
        global is_not_win
        if not is_not_win:
            Beep(frequency, duration)  # Based on Windows API - Single platform support
        else:
            print('\a')  # Cross platform. Limited control over frequency and duration

    # Warn using feedback mechanism - audio
    def warn(self):
        self.attribute.isWarned = True
        self.attribute.isFlushed = False
        self.console_output('Warning Generated!')
        self.beep(frequency=5000, duration=2000)

    # Flush attendance if lecture over, or on program close
    def flush(self):

        # If system isn't authenticated, return execution sequence
        if not self.attribute.isAuthenticated:
            return

        # If host faculty present, starts export procedure
        if self.attribute.host_faculty:

            # Collective dictionary with lecture record
            attendance = dict()

            # Assign copies of dictionary not alias
            attendance['host'] = self.attribute.host_faculty.copy()
            attendance['attendees'] = self.attribute.attendees.copy()

            # Export attendance; return unique subject key
            key = self.export_attendance(attendance)

            # If key returned, add attendance entry to total lecture record of day with key as index
            if key:
                self.attribute.attendance_all[key] = attendance.copy()
            else:
                return

        # Set flags for warning and flush state; clear exported lecture records
        self.attribute.isFlushed = True
        self.attribute.isAuthenticated = False
        self.attribute.host_faculty.clear()
        self.attribute.attendees.clear()
        self.obj.student.student_list.clear()

        # Print message and give feedback
        self.console_output('Attendance Flushed!')
        self.beep(frequency=3500, duration=1000)

    # Adjust parameters and call export functions
    def export_attendance(self, attendance):

        # Fetch subject name from Scheduler class based on batch_name and lecture_number
        # subject = self.obj.scheduler.lecture(self.attribute.batch_name, self.attribute.lecture_number)

        # if subject is None:
        #     print("Error: Subject is not specified.")
        #     return None
    
        # Create shell container to store attendance contents
        attendance_dict = dict()

        # Create unique key by concatenating subject name and lecture time slot of current lecture
        key_name = ' ' +\
                   str(self.attribute.lecture_time).replace("['", "(").replace("']", ")").replace("', '", "-")

        # Add lecture record to whole day's attendance
        attendance_dict[key_name] = attendance.copy()

        # Call export function which returns path of excel_file it created
        excel_file = self.obj.export.export(attendance_dict)

        # Mail attendance record to host faculty's email - pass excel file path as parameter
        self.obj.mailer.send_attendance(attachment=excel_file, email=self.attribute.host_faculty['Email'],
                                        attendees_len=len(self.attribute.attendees))

        # Return unique key
        return key_name

    # Set authentication flags; provide authentication feedback
    def auth(self, faculty_data):
        self.attribute.isAuthenticated = True
        self.attribute.host_faculty = faculty_data
        self.console_output(f"Lecture held by: {faculty_data['Name']} ({faculty_data['Code']})")
        self.beep(frequency=2500, duration=1250)

    # Print text on frame
    def frame_text(self, frame, text, font=cv2.FONT_HERSHEY_PLAIN):
        cv2.putText(
            img=frame,
            text=text,
            org=(5, 30),
            fontFace=font,
            fontScale=1.5,
            color=(255, 0, 0),
            thickness=3)
        
    # Add attendee record; give feedback
    def attend(self, attendee):
        if str(attendee) not in str(self.attribute.attendees):
            self.attribute.attendees.append(attendee)
            self.beep(frequency=2500, duration=300)

    # Disable interface capabilties.
    def college_over(self):
        if not self.obj.timer.lecture_time():
            self.button_monitor.setDisabled(True)
            self.button_monitor.setText(self.qtranslate(self.centralwidget_name, 'Outside STI Calbayog School Hours'))

    # Check weekday and compare against time table.
    def is_holiday(self):

        # Check if section name matches one in database
        for section in self.obj.scheduler.lecture_table:

            # Fetch section data
            if self.attribute.batch_name == section:

                # Check if section doesn't have lecture entries in present day schedule
                # Print message and partially disable interface capabilities
                if not datetime.now().strftime("%A") in str(self.obj.scheduler.lecture_table[section].keys()):
                    self.btn_monitor.setText(self.qtranslate(self.centralwidget_name, 'Outside STI Calbayog Class Hours!!'))
                    for btn in self.buttons:
                        if btn == 'btn_attendee' or btn == 'btn_config':
                            continue
                        getattr(self, btn).setDisabled(True)
                if not datetime.now().strftime("%A") in str(self.obj.scheduler.lecture_table[section].keys()):
                    self.btn_logoutmonitor.setText(self.qtranslate(self.centralwidget_name, 'Outside STI Calbayog Class Hours!!'))
                    for btn in self.buttons:
                        if btn == 'btn_attendee' or btn == 'btn_config':
                            continue
                        getattr(self, btn).setDisabled(True)

                # Break loop to prevent execution of 'else' block
                break

        # No match found for section name - print message and partially disable interface capabilities
        else:
            self.console_output('Section name not found in database!')
            for btn in self.buttons:
                if btn == 'btn_config':
                    continue
                getattr(self, btn).setDisabled(True)


class Monitor(Utility):

    # Start attendance monitor - write authentication check and control here
    def monitor_cam(self):

        # Print message
        self.console_output('Starting login monitor mode...')

        # Setup video capture stream from specified capture device
        self.capture = cv2.VideoCapture(self.capture_device, cv2.CAP_DSHOW)

        # Change START on button_monitor to STOP
        self.button_monitor.setText(self.qtranslate(self.centralwidget_name, 'STOP Login Monitor'))

        # If self.monitor is True, attendance monitor starts

        # Disable window close button
        self.main_window.setWindowFlag(QtCore.Qt.WindowCloseButtonHint, False)

        # Main window closes on close button state change - show window again
        if not self.main_window.isVisible():
            self.main_window.show()

        while self.monitor:

            # Check for lecture time, warning, lecture_flush, college over & total_flush - move this step up sequence
            time_check = self.time_check()
            if time_check == -1:
                return

            # Capture frames from self.capture source and flips them correctly
            _, frame = self.capture.read()
            frame = cv2.flip(frame, 1)

            # Check for qr codes in frame - handle authentication and attendance
            # Return frame with added properties
            process_frame = self.processor(frame)
            if process_frame == -1:
                continue

            # cv2 frames are numpy.ndarray type objects: convert them to QImage
            qimage = QtGui.QImage(frame, frame.shape[1], frame.shape[0], frame.shape[1] * 3,
                                 QtGui.QImage.Format_RGB888).rgbSwapped()

            # Further convert QImage to QPixmap object
            cam_pixmap = QtGui.QPixmap(qimage)

            # Render QPixmap to camera container(QFrame object)
            self.frame_cam.setPixmap(cam_pixmap)

            # # Check if main window is closed: stops attendance monitor if it is, or reopens window in same state
            # if not self.main_window.isVisible():
            #     # self.stop_monitor()
            #     self.main_window.show()

            # Tell event loop to process events
            self.application.processEvents()

        # Remark: Only reached on stop_monitor() call - syncs with total_flush
        else:
            if self.cam_on:
                self.stop_monitor()

    # Stop attendance monitor
    def stop_monitor(self):

        # Print stopping message
        self.console_output('Stopping login monitor mode...')

        # Enable windows close button
        self.main_window.setWindowFlag(QtCore.Qt.WindowCloseButtonHint, True)

        # Main window closes on close button state change - show window again
        if not self.main_window.isVisible():
            self.main_window.show()

        # Reset application state
        self.monitor = False
        self.cam_on = False

        # Release camera lock
        self.capture.release()

        # Clear camera frame - QFrame and reset button_monitor text
        self.frame_cam.clear()
        self.frame_cam.setText(self.qtranslate(self.centralwidget_name, 'Camera Output'))
        self.button_monitor.setText(self.qtranslate(self.centralwidget_name, 'START Login Monitor'))

        # Check if lecture active - flush attendance if True
        if self.attribute.host_faculty:
            self.flush()

        # Export and mail record
        if self.attribute.attendance_all:
            excel_file = self.obj.export.export(self.attribute.attendance_all)
            self.obj.mailer.send_attendance(attachment=excel_file,
                                            lecture_len=len(self.attribute.attendance_all.keys()))

    # Trigger to start/stop monitor cam
    def monitor_trigger(self):

        # Enable/disable buttons
        if self.monitor:
            self.button_monitor.setDisabled(True)
            self.button_logoutmonitor.setDisabled(False)
        else:
            self.button_monitor.setDisabled(False)
            self.button_logoutmonitor.setDisabled(True)

        # Check status of attendance monitor to determine further action
        if self.cam_on:
            self.stop_monitor()

        else:
            # Set application state and launch camera monitor
            self.monitor = True
            self.cam_on = True
            self.monitor_cam()

    # Checks present time on each iteration
    def time_check(self):

        # Time-checks and operations:
        # First perform check for current time vs time slots in time table
        # Then obtain current time window with start and end element
        # On change in lecture_number, reset isWarned attribute; change self.lecture_time
        if self.obj.timer.lecture_time():

            # Obtain current lecture end (datetime object), lecture number, and lecture time slot
            self.attribute.lecture_end, lecture_num, self.attribute.lecture_tuple = self.obj.timer.lecture_time()

            # Evaluate if lecture number has changed
            if not self.attribute.lecture_number == lecture_num:

                # Update lecture time slot on lecture number change; reset isWarned flag
                self.attribute.lecture_time = list(self.attribute.lecture_tuple)
                self.attribute.isWarned = False

            # Update current lecture_number
            self.attribute.lecture_number = lecture_num

        else:

            # If current time doesn't match a time slot, college is over
            self.console_output('Outside STI Calbayog School Hours!')
            self.monitor = False
            self.stop_monitor()
            self.college_over()

            # Return a value to signal the loop
            return -1

        # Update seconds left to warn and flush - based on time slots in lecture table
        # Fire time-based events: warn() and flush()
        warn_time = self.attribute.lecture_end - timedelta(minutes=self.attribute.warning_period_minutes)
        warn_delay = (warn_time - datetime.now()).total_seconds()
        flush_delay = (self.attribute.lecture_end - datetime.now()).total_seconds()
        if not self.attribute.isWarned and (0 <= warn_delay <= 1):
            self.warn()
        elif not self.attribute.isFlushed and (0 <= flush_delay <= 1):
            self.flush()

    # Process frame for QR Code
    def processor(self, frame):

        # If exception thrown, then no camera is present - or invalid source defined
        try:
            decoded_list = decode(frame)
        except TypeError:
            self.console_output('Camera not found!')
            self.monitor = False
            return -1

        # If frame doesn't contain QR Code, return flow of execution
        if not decoded_list:
            return
        else:
            decoded = decoded_list[0]

            # Decoded object is bytes type
            qr_data = decoded.data.decode('utf-8')

            # Authentication tokens are all digits, and have less number of digits than tokenLimit.
            # Also, only check if not authenticated already
            if qr_data.isdigit() and len(qr_data) <= len(str(self.attribute.tokenLimit
                                                             )) and not self.attribute.isAuthenticated:

                # Authenticate faculty
                session_faculty = self.obj.faculty.auth(qr_data)

                # If authenticated successfully, call self.auth() with returned data
                if session_faculty: self.auth(session_faculty)

            # If already authenticated and qrdata is all digits, print 'active session' message
            elif qr_data.isdigit() and self.attribute.isAuthenticated:
                self.frame_text(frame=frame, text='Session Activated!')

            # If authenticated, and input is string, check if attendee is in database
            elif self.attribute.isAuthenticated and not qr_data.isdigit():
                qr_data_list = qr_data.strip('][').replace("'", '').split(', ')

                # Check input data signature. Ignore bad input.
                try:
                    verified_student = self.obj.student.validate(roll=qr_data_list[0], name=qr_data_list[1])
                except IndexError:
                    return

                # If attendee data in database, print message and call attend()
                if verified_student:
                    image_text = qr_data_list[0] + ': ' + qr_data_list[1].replace('  ', ' ')
                    self.frame_text(frame=frame, text=image_text)
                    self.attend(verified_student)

    # Start attendance monitor - write authentication check and control here
    def logoutmonitor_cam(self):

        # Print message 
        self.console_output('Starting logout monitor mode...')

        # Setup video capture stream from specified capture device
        self.capture = cv2.VideoCapture(self.capture_device, cv2.CAP_DSHOW)

        # Change START on button_monitor to STOP
        self.button_logoutmonitor.setText(self.qtranslate(self.centralwidget_name, 'STOP Logout Monitor'))

        # If self.monitor is True, attendance monitor starts

        # Disable window close button
        self.main_window.setWindowFlag(QtCore.Qt.WindowCloseButtonHint, False)

        # Main window closes on close button state change - show window again
        if not self.main_window.isVisible():
            self.main_window.show()

        while self.logoutmonitor:

            # Check for lecture time, warning, lecture_flush, college over & total_flush - move this step up sequence
            time_check = self.time_check()
            if time_check == -1:
                return

            # Capture frames from self.capture source and flips them correctly
            _, frame = self.capture.read()
            frame = cv2.flip(frame, 1)

            # Check for qr codes in frame - handle authentication and attendance
            # Return frame with added properties
            process_frame = self.logprocessor(frame)
            if process_frame == -1:
                continue

            # cv2 frames are numpy.ndarray type objects: convert them to QImage
            qimage = QtGui.QImage(frame, frame.shape[1], frame.shape[0], frame.shape[1] * 3,
                                 QtGui.QImage.Format_RGB888).rgbSwapped()

            # Further convert QImage to QPixmap object
            cam_pixmap = QtGui.QPixmap(qimage)

            # Render QPixmap to camera container(QFrame object)
            self.frame_cam.setPixmap(cam_pixmap)

            # # Check if main window is closed: stops attendance monitor if it is, or reopens window in same state
            # if not self.main_window.isVisible():
            #     # self.stop_monitor()
            #     self.main_window.show()

            # Tell event loop to process events
            self.application.processEvents()

        # Remark: Only reached on stop_monitor() call - syncs with total_flush
        else:
            if self.logoutcam_on:
                self.stop_logoutmonitor()

    # Stop attendance monitor
    def stop_logoutmonitor(self):

        # Print stopping message
        self.console_output('Stopping logout monitor mode...')

        # Enable windows close button
        self.main_window.setWindowFlag(QtCore.Qt.WindowCloseButtonHint, True)

        # Main window closes on close button state change - show window again
        if not self.main_window.isVisible():
            self.main_window.show()

        # Reset application state
        self.logoutmonitor = False
        self.logoutcam_on = False

        # Release camera lock
        self.capture.release()

        # Clear camera frame - QFrame and reset button_monitor text
        self.frame_cam.clear()
        self.frame_cam.setText(self.qtranslate(self.centralwidget_name, 'Camera Output'))
        self.button_logoutmonitor.setText(self.qtranslate(self.centralwidget_name, 'START Logout Monitor'))

        # Check if lecture active - flush attendance if True
        if self.attribute.host_faculty:
            self.flush()

        # Export and mail record
        if self.attribute.attendance_all:
            excel_file = self.obj.logexport.export(self.attribute.attendance_all)
            self.obj.mailer.send_attendance(attachment=excel_file,
                                            lecture_len=len(self.attribute.attendance_all.keys()))

    # Trigger to start/stop monitor cam
    def logoutmonitor_trigger(self):

        # Enable/disable buttons
        if self.logoutmonitor:
            self.button_logoutmonitor.setDisabled(True)
            self.button_monitor.setDisabled(False)
        else:
            self.button_logoutmonitor.setDisabled(False)
            self.button_monitor.setDisabled(True)

        # Check status of attendance monitor to determine further action
        if self.logoutcam_on:
            self.stop_logoutmonitor()

        else:
            # Set application state and launch camera monitor
            self.logoutmonitor = True
            self.logoutcam_on = True
            self.logoutmonitor_cam()

    # Checks present time on each iteration
    def time_check(self):

        # Time-checks and operations:
        # First perform check for current time vs time slots in time table
        # Then obtain current time window with start and end element
        # On change in lecture_number, reset isWarned attribute; change self.lecture_time
        if self.obj.timer.lecture_time():

            # Obtain current lecture end (datetime object), lecture number, and lecture time slot
            self.attribute.lecture_end, lecture_num, self.attribute.lecture_tuple = self.obj.timer.lecture_time()

            # Evaluate if lecture number has changed
            if not self.attribute.lecture_number == lecture_num:

                # Update lecture time slot on lecture number change; reset isWarned flag
                self.attribute.lecture_time = list(self.attribute.lecture_tuple)
                self.attribute.isWarned = False

            # Update current lecture_number
            self.attribute.lecture_number = lecture_num

        else:

            # If current time doesn't match a time slot, college is over
            self.console_output('Outside STI Calbayog School Hours!')
            self.logoutmonitor = False
            self.stop_logoutmonitor()
            self.college_over()

            # Return a value to signal the loop
            return -1

        # Update seconds left to warn and flush - based on time slots in lecture table
        # Fire time-based events: warn() and flush()
        warn_time = self.attribute.lecture_end - timedelta(minutes=self.attribute.warning_period_minutes)
        warn_delay = (warn_time - datetime.now()).total_seconds()
        flush_delay = (self.attribute.lecture_end - datetime.now()).total_seconds()
        if not self.attribute.isWarned and (0 <= warn_delay <= 1):
            self.warn()
        elif not self.attribute.isFlushed and (0 <= flush_delay <= 1):
            self.flush()

    # Process frame for QR Code
    def logprocessor(self, frame):

        # If exception thrown, then no camera is present - or invalid source defined
        try:
            decoded_list = decode(frame)
        except TypeError:
            self.console_output('Camera not found!')
            self.logoutmonitor = False
            return -1

        # If frame doesn't contain QR Code, return flow of execution
        if not decoded_list:
            return
        else:
            decoded = decoded_list[0]

            # Decoded object is bytes type
            qr_data = decoded.data.decode('utf-8')

            # Authentication tokens are all digits, and have less number of digits than tokenLimit.
            # Also, only check if not authenticated already
            if qr_data.isdigit() and len(qr_data) <= len(str(self.attribute.tokenLimit
                                                             )) and not self.attribute.isAuthenticated:

                # Authenticate faculty
                session_faculty = self.obj.faculty.auth(qr_data)

                # If authenticated successfully, call self.auth() with returned data
                if session_faculty: self.auth(session_faculty)

            # If already authenticated and qrdata is all digits, print 'active session' message
            elif qr_data.isdigit() and self.attribute.isAuthenticated:
                self.frame_text(frame=frame, text='Session Activated!')

            # If authenticated, and input is string, check if attendee is in database
            elif self.attribute.isAuthenticated and not qr_data.isdigit():
                qr_data_list = qr_data.strip('][').replace("'", '').split(', ')

                # Check input data signature. Ignore bad input.
                try:
                    verified_student = self.obj.logout.validate(roll=qr_data_list[0], name=qr_data_list[1])
                except IndexError:
                    return

                # If attendee data in database, print message and call attend()
                if verified_student:
                    image_text = qr_data_list[0] + ': ' + qr_data_list[1].replace('  ', ' ')
                    self.frame_text(frame=frame, text=image_text)
                    self.attend(verified_student)


class Application(Monitor):

    def __init__(self):
        # Instantiate PyQt5 application and main window
        self.application = QtWidgets.QApplication([])
        self.main_window = MainWindow()

        # Define interface variables
        self.qtranslate = QtCore.QCoreApplication.translate
        self.qfont = QtGui.QFont()
        self.qicon = QtGui.QIcon()

        self.buttons = {'btn_session': 'Generate QR Code Session Tokens',
                        'btn_attendee': 'Generate Student QR Code',
                        'btn_config': 'Configure PaSCAN Values',
                        'btn_logoutmonitor': 'Start Logout Monitor',# Added logout button
                        'btn_monitor': 'START Login Monitor',}  
        self.centralwidget_name = 'PaSCAN Dashboard'
        self.monitor = bool()
        self.cam_on = bool()
        self.logoutmonitor = bool()
        self.logoutcam_on = bool()

        # Specify video input stream/source
        self.capture_device = 0  # 0 means capture from internal camera device

        # Sequential function calls
        # Interface setup functions
        self.setup_dashboard(self.main_window)
        self.setup_cam()
        self.setup_btn()
        self.attach_btn()

        # Add console_emulation
        self.setup_console()

        # Import attributes and objects
        # Passed self.console_output() which prints output to console emulation in application instead of stdout
        self.obj = Object(console_output=self.console_output, qicon=self.qicon, application_window=self.application)
        self.attribute = self.obj.return_attribute_obj()

        # Trigger setup functions
        self.connect_slots()

        # Check if outside college hours
        # Check for holiday
        self.college_over()
        self.is_holiday()

        # Disable logout button if monitor is running
        if self.monitor:
            self.button_logoutmonitor.setDisabled(True)
        else:
            self.button_logoutmonitor.setDisabled(False)

        # Disable monitor button if logoutmonitor is running
        if self.logoutmonitor:
            self.button_monitor.setDisabled(True)
        else:
            self.button_monitor.setDisabled(False)

        # Show main window
        self.main_window.show()

        # Run event loop
        self.application.exec_()

    # Setup main_window
    def setup_dashboard(self, dashboard_window):

        # Set main window object name
        dashboard_window.setObjectName(self.centralwidget_name)

        # Set window fixed size - prevent maximize-restore
        dashboard_window.setFixedSize(1510, 768)

        # Set window title
        dashboard_window.setWindowTitle(self.qtranslate(self.centralwidget_name, "PaSCAN QR Code Attendance Management System"))

        # Name and path of main window icon file
        self.window_icon_container = 'components'
        self.window_icon_filename = 'PaSCAN Logo.png'
        self.window_icon_filepath = abspath(join(dirname(__name__), self.window_icon_container,
                                                 self.window_icon_filename))

        # Set window icon
        self.qicon.addPixmap(QtGui.QPixmap(self.window_icon_filepath))
        dashboard_window.setWindowIcon(self.qicon)

        # Set dashboard_window as central widget (root) which further contains (encloses) other widgets
        self.centralwidget = QtWidgets.QWidget(dashboard_window)
        dashboard_window.setCentralWidget(self.centralwidget)

    # Setup camera frame
    def setup_cam(self):

        self.frame_cam = QtWidgets.QLabel(self.centralwidget)
        self.frame_cam.setGeometry(QtCore.QRect(20, 20, 1000, 510))

        # Give frame a box enclosure (boundary)
        self.frame_cam.setFrameShape(QtWidgets.QFrame.Box)

        # Add font style properties
        self.qfont.setFamily('Segoi UI')
        self.qfont.setPointSize(20)
        self.frame_cam.setFont(self.qfont)
        self.frame_cam.setCursor(QtGui.QCursor(QtCore.Qt.CrossCursor))
        self.frame_cam.setStyleSheet('background: rgb(0, 40, 100); color: white;')
        self.frame_cam.setAlignment(QtCore.Qt.AlignCenter)
        self.frame_cam.setText(self.qtranslate(self.centralwidget_name, 'Camera Output'))

    # Setup right pane buttons
    def setup_btn(self):

        pass
        # Set measurements: button offset, length, margin
        _btn_offset_x_ = 20
        _btn_height_ = 108
        _btn_offset_margin_ = 25

        # Define font properties
        self.qfont.setFamily('Arial')
        self.qfont.setPointSize(13)

        # Identifies second last button
        last_button_key = list(self.buttons.keys())[-2]

        # Instantiate button objects using self.buttons dict() items
        for btn_name, btn_text in self.buttons.items():

            # Bind button object as an attribute of class instance
            setattr(self, btn_name, QtWidgets.QPushButton(self.centralwidget))

            # Fetch button object from class instance object
            button = getattr(self, btn_name)
            button.setGeometry(QtCore.QRect(1035, _btn_offset_x_, 460, _btn_height_))
            button.setText(self.qtranslate(self.centralwidget_name, btn_text))

            # Set font properties
            button.setFont(self.qfont)

            # Update next button offset\
            _btn_offset_x_ = _btn_offset_x_ + _btn_height_ + _btn_offset_margin_

            # Special case when it's the last button - in grid with PaSCAN console
            if btn_name == last_button_key:
                _btn_height_ = int(1.6*_btn_height_)

        
    # Create shell button instance - use this to bridge static slots with dynamic buttons
    def attach_btn(self):
        self.button_session = self.btn_session
        self.button_attendee = self.btn_attendee
        self.button_config = self.btn_config
        self.button_monitor = self.btn_monitor
        self.button_logoutmonitor = self.btn_logoutmonitor

    # Setup console_emulation in application using QTextBrowser object
    def setup_console(self):
        self.textbrowser_console = QtWidgets.QTextBrowser(self.centralwidget)
        self.textbrowser_console.setGeometry(QtCore.QRect(20, 529, 1000, 194))
        self.textbrowser_console.viewport().setProperty("cursor", QtGui.QCursor(QtCore.Qt.IBeamCursor))
        self.textbrowser_console.setStyleSheet("background: rgb(10, 10, 10); color: white;")
        self.qfont.setPointSize(13)
        self.textbrowser_console.setFont(self.qfont)
        self.textbrowser_console.setPlaceholderText(self.qtranslate(self.centralwidget_name, 'PaSCAN QR Code Attendance Management System Console'))

    # Connect push button slots
    def connect_slots(self):
        pass#didi

        # Attach token_generator() as signal to button_session click event
        self.button_session.clicked.connect(partial(self.obj.token.generate_session, self.application))

        # Attach code_generator() as signal to button_monitor_attendee click event
        self.button_attendee.clicked.connect(partial(self.obj.student.code_generator, self.application))

        # Attach config_manager() as signal to button_config click event
        self.button_config.clicked.connect(partial(self.obj.config.config_manager))

        # Attach monitor_trigger() as signal to button_monitor click event
        self.button_monitor.clicked.connect(partial(self.monitor_trigger))

        self.button_logoutmonitor.clicked.connect(partial(self.logoutmonitor_trigger))



    # Define console_output method
    def console_output(self, msg):
        # Fetch text already in QTextBrowser
        pre_text = self.textbrowser_console.toPlainText()

        # Newline addition only if pre_text exist
        if pre_text:
            self.textbrowser_console.setText(self.qtranslate(self.centralwidget_name, (pre_text + '\n' + str(msg))))
        else:
            self.textbrowser_console.setText(self.qtranslate(self.centralwidget_name, str(msg)))

        # Move cursor view to the end - scroll view
        self.textbrowser_console.moveCursor(QtGui.QTextCursor.End)

        # Repaint to force event loop to render at each call
        self.textbrowser_console.repaint()

class MainWindow(QtWidgets.QMainWindow):
    

    def closeEvent(self, event):

        config = ConfigParser()
        config.read(r'C:/Users/Sample file path directory/PaSCAN/config/config.ini')  # Provide the path to your config.ini if it's not in the current directory

        
        batch_name = config.get('values', 'batch_name')

        self.present = str(datetime.now().year) + '-' + str(datetime.now().month) + '-' + str(datetime.now().day)
        self.currentDateTime = str(datetime.now().year) + '-' + str(datetime.now().month) + '-' + str(datetime.now().day) + ' ' + str(datetime.now().strftime("%H-%M-%S"))
        self.loginDateCreated = 'LogIn'
        self.logoutDateCreated = 'Logout'

        forLogin = f"{batch_name} {self.present} {self.loginDateCreated}.xlsx"
        forLogout = f"{batch_name} {self.present} {self.logoutDateCreated}.xlsx"
        file_path_for_Data = r"C:/Users/Sample file path directory/Try PaSCAN/Try PaSCAN/database/attendance/"


        output_file = os.path.join(file_path_for_Data, f"{batch_name} Complete Attendance {self.present}.xlsx")
        fileLogIn = os.path.join(file_path_for_Data, forLogin)
        fileLogOut = os.path.join(file_path_for_Data, forLogout)
        

        try:
            # Read the data from both files
            df1 = pd.read_excel(fileLogIn, header=None)
            df2 = pd.read_excel(fileLogOut, header=None)

            # Create an empty DataFrame with the same number of columns as df1
            empty_df = pd.DataFrame([[''] * df1.shape[1]] * 3, columns=df1.columns)

            # Combine the data vertically with an empty row in between
            combined_df = pd.concat([df1, empty_df, df2], ignore_index=True)

            # Write the combined data to a new Excel file
            combined_df.to_excel(output_file, index=False, header=False)

            # Load the workbook and select the active sheet
            wb = load_workbook(output_file)
            ws = wb.active

            # Set cell width to 35
            for col in ws.columns:
                max_length = 35
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max_length

            # Set alignment and font for all cells
            alignment = Alignment(horizontal='center', vertical='center')
            font = Font(bold=True)

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = alignment
                    cell.font = font

            # Save the workbook
            wb.save(output_file)

            # Create a new directory for the current date
            new_folder_path = os.path.join(file_path_for_Data, f"{batch_name} {self.currentDateTime}")
            os.makedirs(new_folder_path, exist_ok=True)

            # Move the files to the new directory
            shutil.move(fileLogIn, os.path.join(new_folder_path, forLogin))
            shutil.move(fileLogOut, os.path.join(new_folder_path, forLogout))
            shutil.move(output_file, os.path.join(new_folder_path, f"{batch_name} Complete Attendance {self.present}.xlsx"))

        except Exception as e:
            print(f"An error occurred: {e}")


        event.accept()

# Process command line arguments
def argparser():

    # Initiate ArgumentParser(); add argument
    parser = ArgumentParser(prog='PaSCAN', add_help=False,
                            description='PaSCAN QR Code Attendance Management System',
                            epilog='BATA LINK DIDI')
    parser.add_argument('--help', '-H', action='help', help="Show this help message and exit.")
    parser.add_argument('--version', '-V', action='version', version=f'%(prog)s {__version__}',
                        help="Show program's version number and exit.")

    # Check if unknown argument given.
    known_args, unknown_args = parser.parse_known_args()
    if unknown_args: parser.print_help()

    # Return valid argument(s)
    return known_args

if __name__ == '__main__':
    # Process (if passed) command line argument(s)
    args = argparser()

    # Execute GUI from this entry point
    application = Application()

